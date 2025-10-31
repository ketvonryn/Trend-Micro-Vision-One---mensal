import pandas as pd
from openpyxl import load_workbook

def atualiza_aba(arquivo_excel, aba, dados, colunas_adicionais=None):
   
    """
    Atualiza uma aba preexistente do Excel.
    Sempre em modo append — nunca sobrescreve.
    """

    # --- Validação inicial ---
    if not isinstance(dados, pd.DataFrame):
        print(dados)
        return

    # --- Verificação de limite de linhas ---
    total_linhas_df = len(dados)
    workbook = load_workbook(arquivo_excel)

    # Verifica se a aba existe e se não excede a quantidade de linhas permitidas
    if aba not in workbook.sheetnames:
        start_row = 0
    else:
        start_row = workbook[aba].max_row
        if start_row + total_linhas_df > 1_048_576:
            return (
                f"[ERRO] A aba '{aba}' não pode receber {total_linhas_df} novas linhas — "
                f"isso excederia o limite de 1.048.576 linhas do Excel."
            )

    # --- Adiciona colunas extras, se houver ---
    if colunas_adicionais:
        for pos, (col, val) in enumerate(colunas_adicionais):
            dados.insert(loc=pos, column=col, value=val)

    # --- Escrita no Excel ---
    try:
        with pd.ExcelWriter(arquivo_excel, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # header apenas se for uma aba nova
            # Escreve a partir da última linha livre
            dados.to_excel(
                            writer,
                            sheet_name=aba,
                            index=False,
                            header=True,
                            startrow=0
                            )

        print(f"Aba '{aba}' atualizada com sucesso — {total_linhas_df} novas linhas adicionadas.")

    except Exception as e:
        print(f"[ERRO] Falha ao atualizar a aba '{aba}': {e}")
